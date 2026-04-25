import asyncio
import re
import json
import urllib.request
import openpyxl
from datetime import datetime
from playwright.async_api import async_playwright

# ── Supabase ──────────────────────────────────────────────────────────────────
import os
SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://cjxkhvkvhgnlnviykoad.supabase.co")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImNqeGtodmt2aGdubG52aXlrb2FkIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NzY1ODIyMDYsImV4cCI6MjA5MjE1ODIwNn0.eCg-JzEshidI-l7pVsumO_SsXbDOh_s--zvH1jc78g0")
DB_HEADERS = {
    "apikey":        SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type":  "application/json",
    "Prefer":        "return=minimal",
}

results = []


# ── Database ───────────────────────────────────────────────────────────────────
def save_to_db(dealer, product, price_str, url, status="OK"):
    try:
        price_num = float(price_str.replace("$","").replace(",","")) \
                    if status == "OK" else None
        payload = json.dumps({
            "dealer":    dealer,
            "product":   product,
            "buy_price": price_num,
            "url":       url,
            "status":    status,
        }).encode("utf-8")
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/prices",
            data=payload,
            headers=DB_HEADERS,
            method="POST",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.status in (200, 201)
    except Exception as e:
        print(f"    [DB ERROR] {e}")
        return False


# ── Parse price string → float ─────────────────────────────────────────────────
def parse_price(raw):
    """Extract a clean float from a raw price string like '$6,812.00' or 'AUD 6812'."""
    if not raw:
        return None
    cleaned = re.sub(r'[^\d.]', '', raw.replace(",", ""))
    try:
        val = float(cleaned)
        return val if val > 0 else None
    except:
        return None


# ── CSS selector price extraction ─────────────────────────────────────────────
async def get_price_by_selector(page, selectors):
    """
    Try each CSS selector in order.
    Returns the first valid price string found, or None.
    """
    for selector in selectors:
        try:
            el = await page.query_selector(selector)
            if el:
                text = await el.inner_text()
                text = text.strip()
                if text:
                    return text
        except:
            continue
    return None


# ── Regex fallback (whole page) ────────────────────────────────────────────────
def regex_fallback(text, min_aud, max_aud):
    """
    Last resort — scan whole page text for prices in AUD range.
    Only used if CSS selectors find nothing.
    """
    matches = []
    # $6,812.00 or A$6812
    matches += re.findall(r'A?\$\s*([\d]{1,3}(?:,[\d]{3})*(?:\.[\d]{2})?)', text)
    # 6751.58/oz  ← ABC format
    matches += re.findall(r'([\d]{4,6}\.[\d]{2})/oz', text)
    # bare comma number
    matches += re.findall(r'\b([\d]{1,2},[\d]{3}(?:\.[\d]{2})?)\b', text)

    from collections import Counter
    prices = []
    for m in matches:
        try:
            val = float(m.replace(",",""))
            if min_aud <= val <= max_aud:
                prices.append(val)
        except:
            pass

    if not prices:
        return None

    from collections import Counter
    most_common = Counter(prices).most_common(1)[0][0]
    return f"${most_common:,.2f}"


# ── PRODUCT & DEALER DEFINITIONS ──────────────────────────────────────────────
#
# Each dealer entry has:
#   url       — exact product page URL
#   selectors — CSS selectors to try IN ORDER (most specific first)
#   wait      — ms to wait after page load for JS to render
#   min_aud   — minimum valid price in AUD (sanity check)
#   max_aud   — maximum valid price in AUD (sanity check)
#
# Selectors are tried top to bottom. First match wins.
# If all selectors fail, regex_fallback is used as last resort.
#
# Common selectors across AU bullion sites:
#   WooCommerce: .woocommerce-Price-amount, .price ins .amount, .price .amount
#   Generic:     [itemprop="price"], .product-price, .entry-price
# ─────────────────────────────────────────────────────────────────────────────

PRODUCTS = [

    # ══════════════════════════════════════════════════════════
    #  1oz GOLD KANGAROO 2026
    # ══════════════════════════════════════════════════════════
    {
        "name":    "1oz Gold Kangaroo 2026",
        "min_aud": 6000,
        "max_aud": 15000,
        "dealers": [
            {
                "name": "KJC Bullion",
                "url":  "https://www.kjc-gold-silver-bullion.com.au/PD/1-oz-2026-australian-kangaroo-gold-bullion-coin/3003878",
                "wait": 3000,
                "selectors": [
                    ".product-price .price",
                    ".price",
                    "[class*='price']",
                    ".amount",
                ],
            },
            {
                "name": "Perth Mint",
                "url":  "https://www.perthmint.com/shop/bullion/bullion-coins/australian-kangaroo-2026-1oz-gold-bullion-coin/",
                "wait": 7000,
                "selectors": [
                    "[class*='ProductPrice']",
                    "[class*='product-price']",
                    "[class*='Price']",
                    ".price",
                    "[data-testid*='price']",
                ],
            },
            {
                "name": "ABC Bullion",
                "url":  "https://www.abcbullion.com.au/store/Bullion-Coins/gn011oz-perth-mint-kangaroo-gold-coin-9999",
                "wait": 5000,
                "networkidle": True,
                "selectors": [
                    ".final-price",
                    ".price-box .price",
                    "[class*='final-price']",
                    "[class*='product-price']",
                    ".price",
                ],
                # ABC shows spot in /oz format — regex fallback needed
                "use_regex_fallback": True,
            },
            {
                "name": "Ainslie Bullion",
                "url":  "https://ainsliebullion.com.au/Buy/View/Product/Name/1oz-Gold-Coin-2026-Kangaroo-Perth-Mint/ID/673",
                "wait": 4000,
                "selectors": [
                    ".ProductPrice",
                    ".product-price",
                    "[class*='Price']",
                    ".price",
                    "span[class*='price']",
                ],
            },
            {
                "name": "Gold Stackers",
                "url":  "https://www.goldstackers.com.au/product/australian-kangaroo-2026-1-oz-gold-bullion-coin/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Jaggards",
                "url":  "https://www.jaggards.com.au/product/2026-1oz-perth-mint-gold-kangaroo-coin/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Swan Bullion",
                "url":  "https://swanbullion.com/2026-australian-kangaroo-1oz-gold-coin/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Guardian Gold",
                "url":  "https://guardian-gold.com.au/product/1oz-gold-kang-coin-2026/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
        ],
    },

    # ══════════════════════════════════════════════════════════
    #  1oz SILVER KANGAROO 2026
    # ══════════════════════════════════════════════════════════
    {
        "name":    "1oz Silver Kangaroo 2026",
        "min_aud": 80,
        "max_aud": 300,
        "dealers": [
            {
                "name": "KJC Bullion",
                "url":  "https://www.kjc-gold-silver-bullion.com.au/PD/1-oz-2026-australian-kangaroo-silver-bullion-coin/3003876",
                "wait": 3000,
                "selectors": [
                    ".product-price .price",
                    ".price",
                    "[class*='price']",
                    ".amount",
                ],
            },
            {
                "name": "Perth Mint",
                "url":  "https://www.perthmint.com/shop/bullion/bullion-coins/australian-kangaroo-2026-1oz-silver-bullion-coin-in-pouch/",
                "wait": 12000,
                "selectors": [
                    "[class*='ProductPrice']",
                    "[class*='product-price']",
                    "[class*='Price']",
                    ".price",
                    "[data-testid*='price']",
                ],
            },
            {
                "name": "ABC Bullion",
                "url":  "https://www.abcbullion.com.au/store/Bullion-Coins/silver-coins",
                "wait": 5000,
                "networkidle": True,
                "selectors": [
                    ".final-price",
                    ".price-box .price",
                    "[class*='final-price']",
                    ".price",
                ],
                "use_regex_fallback": True,
            },
            {
                "name": "Ainslie Bullion",
                "url":  "https://ainsliebullion.com.au/Buy/View/Product/Name/1oz-Silver-Coin-2026-Kangaroo-Perth-Mint/ID/677",
                "wait": 4000,
                "selectors": [
                    ".ProductPrice",
                    ".product-price",
                    "[class*='Price']",
                    ".price",
                ],
            },
            {
                "name": "Gold Stackers",
                "url":  "https://www.goldstackers.com.au/product/perth-mint-2026-kangaroo-silver-coin-1-oz/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Jaggards",
                "url":  "https://www.jaggards.com.au/product/2026-1oz-perth-mint-silver-kangaroo-coin/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Swan Bullion",
                "url":  "https://swanbullion.com/2026-australian-kangaroo-1oz-silver-coin/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Guardian Gold",
                "url":  "https://guardian-gold.com.au/product/1oz-silver-kangaroo-coin-2026/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
        ],
    },

    # ══════════════════════════════════════════════════════════
    #  1oz GOLD BAR (Perth Mint Minted)
    # ══════════════════════════════════════════════════════════
    {
        "name":    "1oz Gold Bar (Perth Mint)",
        "min_aud": 6000,
        "max_aud": 15000,
        "dealers": [
            {
                "name": "KJC Bullion",
                "url":  "https://www.kjc-gold-silver-bullion.com.au/CT/perth-mint-gold-bars-1oz-gold/220241/1",
                "wait": 7000,
                "selectors": [
                    ".product-price .price",
                    ".price",
                    "[class*='price']",
                    ".amount",
                ],
            },
            {
                "name": "Perth Mint",
                "url":  "https://www.perthmint.com/shop/bullion/cast-bars/perth-mint-1oz-gold-cast-bar/",
                "wait": 12000,
                "selectors": [
                    "[class*='ProductPrice']",
                    "[class*='product-price']",
                    "[class*='Price']",
                    ".price",
                    "[data-testid*='price']",
                ],
            },
            {
                "name": "ABC Bullion",
                "url":  "https://www.abcbullion.com.au/store/gold/gabg011oz-abc-gold-cast-bar-9999",
                "wait": 5000,
                "networkidle": True,
                "selectors": [
                    ".final-price",
                    ".price-box .price",
                    "[class*='final-price']",
                    ".price",
                ],
                "use_regex_fallback": True,
            },
            {
                "name": "Ainslie Bullion",
                "url":  "https://ainsliebullion.com.au/Buy/View/Product/Name/1oz-Perth-Mint-Gold-Cast-Bar/ID/32",
                "wait": 4000,
                "selectors": [
                    ".ProductPrice",
                    ".product-price",
                    "[class*='Price']",
                    ".price",
                ],
            },
            {
                "name": "Gold Stackers",
                "url":  "https://www.goldstackers.com.au/product/perth-mint-cast-gold-bar-1-oz/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Jaggards",
                "url":  "https://www.jaggards.com.au/product/1oz-perth-mint-gold-minted-bar/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Swan Bullion",
                "url":  "https://swanbullion.com/perth-mint-1oz-gold-minted-bar/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
            {
                "name": "Guardian Gold",
                "url":  "https://guardian-gold.com.au/product/1oz-perth-mint-gold-cast-bar/",
                "wait": 4000,
                "selectors": [
                    ".woocommerce-Price-amount",
                    ".price ins .woocommerce-Price-amount",
                    ".price .amount",
                    ".price",
                ],
            },
        ],
    },

    # ══════════════════════════════════════════════════════════
    #  1g GOLD MINTED BAR (Perth Mint)
    # ══════════════════════════════════════════════════════════
    {
        "name":    "1g Gold Minted Bar (Perth Mint)",
        "min_aud": 200,
        "max_aud": 600,
        "dealers": [
            {
                "name": "KJC Bullion",
                "url":  "https://www.kjc-gold-silver-bullion.com.au/PD/1-g-perth-mint-gold-bullion-minted-bar/2417",
                "wait": 3000,
                "selectors": [".product-price .price", ".price", "[class*=\'price\']"],
            },
            {
                "name": "Perth Mint",
                "url":  "https://www.perthmint.com/shop/bullion/minted-bars/kangaroo-1g-minted-gold-bar/",
                "wait": 9000,
                "selectors": ["[class*=\'ProductPrice\']","[class*=\'product-price\']","[class*=\'Price\']",".price"],
            },
            {
                "name": "ABC Bullion",
                "url":  "https://www.abcbullion.com.au/store/gold/abc-bullion-gold",
                "wait": 5000,
                "networkidle": True,
                "selectors": [".final-price",".price-box .price","[class*=\'final-price\']",".price"],
                "use_regex_fallback": True,
            },
            {
                "name": "Ainslie Bullion",
                "url":  "https://ainsliebullion.com.au/Buy/View/Product/Name/1g-minted-gold-bar-Perth-Mint/ID/25",
                "wait": 4000,
                "selectors": [".ProductPrice",".product-price","[class*=\'Price\']",".price"],
            },
            {
                "name": "Gold Stackers",
                "url":  "https://www.goldstackers.com.au/product/perth-mint-kangaroo-gold-bar-1g/",
                "wait": 4000,
                "selectors": [".woocommerce-Price-amount",".price ins .woocommerce-Price-amount",".price .amount",".price"],
            },
            {
                "name": "Jaggards",
                "url":  "https://www.jaggards.com.au/product/1g-perth-mint-gold-minted-bar/",
                "wait": 4000,
                "selectors": [".woocommerce-Price-amount",".price ins .woocommerce-Price-amount",".price .amount",".price"],
            },
            {
                "name": "Swan Bullion",
                "url":  "https://swanbullion.com/perth-mint-1g-gold-minted-bar/",
                "wait": 4000,
                "selectors": [".woocommerce-Price-amount",".price ins .woocommerce-Price-amount",".price .amount",".price"],
            },
            {
                "name": "Guardian Gold",
                "url":  "https://guardian-gold.com.au/product/1g-perth-mint-gold-minted-bar/",
                "wait": 4000,
                "selectors": [".woocommerce-Price-amount",".price ins .woocommerce-Price-amount",".price .amount",".price"],
            },
        ],
    },
]


# ── Core scrape function ───────────────────────────────────────────────────────
async def scrape_dealer(page, dealer, product_name, min_aud, max_aud):
    name = dealer["name"]
    url  = dealer["url"]

    try:
        await page.goto(url, timeout=60000, wait_until="domcontentloaded")
        if dealer.get("networkidle"):
            try:
                await page.wait_for_load_state("networkidle", timeout=12000)
            except:
                pass
        await page.wait_for_timeout(dealer.get("wait", 3000))

        price_str = None
        method    = "?"

        # 1. Try CSS selectors first
        raw = await get_price_by_selector(page, dealer["selectors"])
        if raw:
            val = parse_price(raw)
            if val and min_aud <= val <= max_aud:
                price_str = f"${val:,.2f}"
                method    = "CSS"

        # 2. Fall back to regex on full page text if needed
        if not price_str:
            text  = await page.inner_text("body")
            price_str = regex_fallback(text, min_aud, max_aud)
            if price_str:
                method = "regex"

        if price_str:
            results.append({
                "product": product_name, "dealer": name,
                "price": price_str, "url": url, "status": "OK",
            })
            saved = save_to_db(name, product_name, price_str, url, "OK")
            db_ok = "✓ db" if saved else "✗ db"
            print(f"  ✓  {name:28s} {price_str:>12s}  [{method}] [{db_ok}]")
        else:
            try:
                text    = await page.inner_text("body")
                preview = " ".join(text.split())[:250]
                print(f"  ✗  {name:28s} NOT FOUND")
                print(f"       URL     : {url}")
                print(f"       Preview : {preview}")
            except:
                print(f"  ✗  {name:28s} NOT FOUND (page unreadable)")
            results.append({
                "product": product_name, "dealer": name,
                "price": "NOT FOUND", "url": url, "status": "NOT FOUND",
            })
            save_to_db(name, product_name, "NOT FOUND", url, "NOT FOUND")

    except Exception as e:
        err = str(e)[:80]
        results.append({
            "product": product_name, "dealer": name,
            "price": "ERROR", "url": url, "status": err,
        })
        save_to_db(name, product_name, "ERROR", url, err)
        print(f"  ✗  {name:28s} {'ERROR':>12s}  {err}")


# ── Excel export ───────────────────────────────────────────────────────────────
def export_excel():
    wb    = openpyxl.Workbook()
    first = True
    hdr_font  = openpyxl.styles.Font(bold=True, color="FFFFFF")
    hdr_fill  = openpyxl.styles.PatternFill("solid", fgColor="1A1A14")
    ok_fill   = openpyxl.styles.PatternFill("solid", fgColor="E8F5EE")
    err_fill  = openpyxl.styles.PatternFill("solid", fgColor="FDE8E8")
    gold_font = openpyxl.styles.Font(bold=True, color="8B6914", size=12)

    for product in PRODUCTS:
        pname    = product["name"]
        rows     = [r for r in results if r["product"] == pname]
        ok_rows  = sorted(
            [r for r in rows if r["status"] == "OK"],
            key=lambda r: float(r["price"].replace("$","").replace(",",""))
        )
        err_rows = [r for r in rows if r["status"] != "OK"]

        ws = wb.active if first else wb.create_sheet()
        ws.title = pname[:31]
        first = False

        for col, h in enumerate(
            ["Rank","Dealer","Buy Price (AUD)","URL","Status","Scraped At"], 1
        ):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill

        for row, r in enumerate(ok_rows + err_rows, 2):
            rank = row - 1 if r["status"] == "OK" else "–"
            ws.cell(row=row, column=1, value=rank)
            ws.cell(row=row, column=2, value=r["dealer"])
            ws.cell(row=row, column=3, value=r["price"])
            ws.cell(row=row, column=4, value=r["url"])
            ws.cell(row=row, column=5, value=r["status"])
            ws.cell(row=row, column=6,
                    value=datetime.now().strftime("%Y-%m-%d %H:%M"))
            fill = ok_fill if r["status"] == "OK" else err_fill
            for col in range(1, 7):
                ws.cell(row=row, column=col).fill = fill
            if row == 2 and r["status"] == "OK":
                ws.cell(row=row, column=3).font = gold_font

        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 75
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 18

    filename = f"bullion_prices_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(filename)
    print(f"\n  ✓ Excel saved → {filename}")


# ── Main ───────────────────────────────────────────────────────────────────────
async def main():
    print("=" * 65)
    print("  GoldSilverPrices.com.au — CSS Selector Scraper v3")
    print("  Runs every 8 hours via Task Scheduler")
    print("=" * 65)
    print(f"  Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    # Test DB
    print("  Testing database connection...")
    try:
        req = urllib.request.Request(
            f"{SUPABASE_URL}/rest/v1/prices?select=id&limit=1",
            headers=DB_HEADERS, method="GET",
        )
        with urllib.request.urlopen(req, timeout=10) as resp:
            resp.read()
        print("  ✓ Supabase connected\n")
    except Exception as e:
        print(f"  ✗ DB failed: {e}")
        return

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=True,
            args=["--no-sandbox","--disable-blink-features=AutomationControlled"]
        )
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            viewport={"width": 1280, "height": 900},
            extra_http_headers={
                "Accept-Language": "en-AU,en;q=0.9",
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
            }
        )
        page = await context.new_page()

        for product in PRODUCTS:
            pname   = product["name"]
            min_aud = product["min_aud"]
            max_aud = product["max_aud"]
            print(f"\n{'─'*65}")
            print(f"  📦 {pname}  (AUD ${min_aud:,}–${max_aud:,})")
            print(f"{'─'*65}")
            for dealer in product["dealers"]:
                await scrape_dealer(page, dealer, pname, min_aud, max_aud)

        await browser.close()

    # Summary
    print(f"\n{'='*65}")
    print("  RESULTS")
    print(f"{'='*65}")
    total_ok = total_all = 0
    for product in PRODUCTS:
        pname = product["name"]
        ok    = [r for r in results
                 if r["product"] == pname and r["status"] == "OK"]
        total = len(product["dealers"])
        total_ok  += len(ok)
        total_all += total
        sorted_ok = sorted(
            ok, key=lambda r: float(r["price"].replace("$","").replace(",",""))
        )
        print(f"\n  {pname}")
        for i, r in enumerate(sorted_ok, 1):
            mark = " ← cheapest" if i == 1 else ""
            print(f"    {i}. {r['dealer']:28s} {r['price']:>12s}{mark}")
        err = [r for r in results
               if r["product"] == pname and r["status"] != "OK"]
        if err:
            print(f"    ✗ Missed: {', '.join(r['dealer'] for r in err)}")
        print(f"    ✓ {len(ok)}/{total} dealers")

    print(f"\n  Total captured: {total_ok}/{total_all}")
    print(f"{'='*65}")
    export_excel()
    print("="*65)


if __name__ == "__main__":
    asyncio.run(main())